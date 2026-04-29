"""
Excel Extraction Service
------------------------
Uses openpyxl to read all sheets and convert to a structured text representation
suitable for Claude API processing.
"""
import logging
from typing import Optional

logger = logging.getLogger(__name__)


def extract_text_from_excel(file_path: str) -> str:
    """
    Read all sheets from an Excel file and return a structured text representation.
    """
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_texts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text: list[str] = []
            has_content = False

            for row in ws.iter_rows(values_only=True):
                # Skip entirely empty rows
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                has_content = True
                cells = []
                for cell in row:
                    if cell is None:
                        cells.append("")
                    else:
                        # Format numbers clearly
                        if isinstance(cell, float) and cell == int(cell):
                            cells.append(str(int(cell)))
                        else:
                            cells.append(str(cell))
                rows_text.append(" | ".join(cells))

            if has_content:
                sheet_texts.append(
                    f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows_text)
                )

        wb.close()
        full_text = "\n\n".join(sheet_texts)
        logger.info(
            "Excel extracted: %d sheets, %d chars", len(sheet_texts), len(full_text)
        )
        return full_text

    except Exception as exc:
        logger.error("Excel extraction failed: %s", exc)
        raise RuntimeError(f"Excel extraction failed: {exc}") from exc
